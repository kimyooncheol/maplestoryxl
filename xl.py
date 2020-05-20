# -*- encoding: utf-8 -*-
#-*- coding:utf-8 -*-
from datetime import datetime
import openpyxl
import f_get
import re
from openpyxl.styles import PatternFill, Color
import numpy as np

#엑셀 저장
def xlsx_Save() :
    now = datetime.now()
    wb.save(now.strftime('%Y-%m-%d')+'.xlsx')
#유저 정보 받기
def get_UserInfo(htmls):
    userLevel = []
    userJob = []
    userFloor = []
    userDate = []
    for html in htmls :
        userLevel.append(f_get.get_UserLevel(html))
        userJob.append(f_get.get_UserJob(html))
        userFloor.append(f_get.get_UserFloor(html))
        userDate.append(f_get.get_UserDate(html))
    return userJob,userLevel,userFloor,userDate

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = '정보'
sheet['A1'] = '닉네임'
sheet['B1'] = '직업'
sheet['C1'] = '레벨'
sheet['D1'] = '무릉'
sheet['E1'] = '마지막 접속일'
sheet['F1'] = '보스'
sheet.column_dimensions['A'].width = 13
sheet.column_dimensions['B'].width = 15
sheet.column_dimensions['E'].width = 14
font = openpyxl.styles.fonts.Font(bold=True)
sheet['A1'].font = font
sheet['B1'].font = font
sheet['C1'].font = font
sheet['D1'].font = font
sheet['E1'].font = font
sheet['F1'].font = font
sheet['A1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['B1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['C1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['D1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['E1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['F1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
sheet['A1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
sheet['B1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
sheet['C1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
sheet['D1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
sheet['E1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
sheet['F1'].fill = PatternFill(patternType='solid', fgColor=Color('848484'))
guildName, guildWorld = f_get.get_Guild()
userNames = np.concatenate((f_get.get_Admin_Username(guildName,f_get.f_input_world(guildWorld)),f_get.get_Username(guildName,f_get.f_input_world(guildWorld))),axis=None)
print(userNames)
htmls = f_get.get_UserInfohtml(userNames)
for a in range(len(userNames)) :
    sheet.cell(row=a+2, column=1).value = userNames[a]
    sheet.cell(row=a+2, column=1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
userJob,userLevel,userFloor,userDate = get_UserInfo(htmls)


def rep(lists) :
    rep1 = str(lists)
    rep2 = rep1.replace("[","")
    rep3 = rep2.replace("]","")
    rep4 = rep3.replace("'","")
    if re.findall('\d+', rep4):
        return(int(re.findall('\d+', rep4)[0]))
    return rep4

for a in range(len(userNames)) :
    sheet.cell(row=a+2, column=2).value = rep(userJob[a])
    sheet.cell(row=a+2, column=2).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=a+2, column=3).value = rep(userLevel[a])
    sheet.cell(row=a+2, column=3).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=a+2, column=4).value = rep(userFloor[a])
    sheet.cell(row=a+2, column=4).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=a+2, column=5).value = rep(userDate[a])
    sheet.cell(row=a+2, column=5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    sheet.cell(row=a+2, column=6).value = '=IF(D%d="X","X",IF(D%d>=46,"노말윌",IF(D%d>=45,"루시드",IF(D%d>=43,"이지루시드",IF(D%d>=41,"스데미","X")))))'%(a+2,a+2,a+2,a+2,a+2)
    sheet.cell(row=a+2, column=6).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    
    if a%2==1:
        for i in range(1,7):
            sheet.cell(row=a+2, column=i).fill = PatternFill(patternType='solid', fgColor=Color('D8D8D8'))
        
xlsx_Save()
